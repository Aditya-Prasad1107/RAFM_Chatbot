"""
Field extraction from folder structure with caching support.
Properly tracks which filename each expression comes from.
"""
import re
import logging
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Union, Optional

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from src.cache import get_cache, ExcelCache

# Configure logger
logger = logging.getLogger(__name__)

# Type aliases for readability
FieldKey = Tuple[str, str]
MappingDict = Dict[FieldKey, List[str]]
ExtractionResult = Tuple[MappingDict, MappingDict, MappingDict]

# Constants
SKIP_VALUES = frozenset({"-", "NA", "N/A"})
EXCEL_ERROR_PATTERNS = (
    '#ref!', '#n/a', '#value!', '#name?', '#null!', '#div/0!'
)


@dataclass(frozen=True)
class ExtractionParams:
    """Immutable extraction parameters for cache key generation."""
    left_col: Union[str, int]
    field_col: Union[str, int]
    value_col: Union[str, int]
    sheet: Union[str, int]
    header_rows_to_skip: int
    strip_key_parts: bool

    def to_dict(self) -> dict:
        """Convert to dictionary for cache compatibility."""
        return {
            'left_col': self.left_col,
            'field_col': self.field_col,
            'value_col': self.value_col,
            'sheet': self.sheet,
            'header_rows_to_skip': self.header_rows_to_skip,
            'strip_key_parts': self.strip_key_parts
        }


@dataclass(frozen=True)
class DynamicExtractionParams:
    """Immutable extraction parameters for dynamic column detection."""
    left_col: Union[str, int]
    field_col: Union[str, int]
    possible_value_cols: Tuple[str, ...]  # Tuple for hashability
    sheet: Union[str, int]
    header_rows_to_skip: int
    strip_key_parts: bool
    header_search_rows: int
    match_text: str

    def to_dict(self) -> dict:
        """Convert to dictionary for cache compatibility."""
        return {
            'left_col': self.left_col,
            'field_col': self.field_col,
            'possible_value_cols': list(self.possible_value_cols),
            'sheet': self.sheet,
            'header_rows_to_skip': self.header_rows_to_skip,
            'strip_key_parts': self.strip_key_parts,
            'header_search_rows': self.header_search_rows,
            'match_text': self.match_text
        }

@dataclass
class FileExtractionResult:
    """Result from processing a single file."""
    mappings: MappingDict
    filename_info: MappingDict
    expression_filenames: MappingDict

    @classmethod
    def from_tuple(cls, data: tuple) -> 'FileExtractionResult':
        """Create from cached tuple format."""
        return cls(
            mappings=data[0],
            filename_info=data[1],
            expression_filenames=data[2]
        )

    def to_tuple(self) -> tuple:
        """Convert to tuple for cache storage."""
        return (self.mappings, self.filename_info, self.expression_filenames)


@dataclass
class CacheStats:
    """Track cache performance statistics."""
    hits: int = 0
    misses: int = 0
    files_processed: int = 0

    def record_hit(self) -> None:
        self.hits += 1

    def record_miss(self) -> None:
        self.misses += 1

    def record_processed(self) -> None:
        self.files_processed += 1

    @property
    def hit_rate(self) -> float:
        total = self.hits + self.misses
        return (self.hits / total * 100) if total > 0 else 0.0

    def log_stats(self) -> None:
        logger.info(
            f"Cache Performance: Hits={self.hits}, Misses={self.misses}, "
            f"Processed={self.files_processed}, Hit Rate={self.hit_rate:.1f}%"
        )


def _col_to_index(col: Union[str, int]) -> int:
    """Convert Excel column letters (A, B, C...) to 1-based numeric indices."""
    if isinstance(col, int):
        if col <= 0:
            raise ValueError(
                "Column must be a positive integer or Excel column letter (A, B, C...)."
            )
        return col

    if not isinstance(col, str):
        raise ValueError(
            "Column must be a positive integer or Excel column letter (A, B, C...)."
        )

    num = 0
    for ch in col.upper():
        if not ('A' <= ch <= 'Z'):
            raise ValueError(f"Invalid column letter: {col}")
        num = num * 26 + (ord(ch) - ord('A') + 1)
    return num


def _normalize_value(cell_value, strip: bool) -> str:
    """Normalize and safely convert cell values."""
    if cell_value is None:
        return ""
    v_str = str(cell_value)
    return v_str.strip() if strip else v_str


def _is_invalid_value(value: str) -> bool:
    """Check if value should be skipped (empty, placeholder, or Excel error)."""
    if not value:
        return True

    if value.strip() in SKIP_VALUES:
        return True

    value_lower = value.lower()
    return any(pattern in value_lower for pattern in EXCEL_ERROR_PATTERNS)


def _get_excel_files(folder_path: Path) -> List[Path]:
    """Get list of valid Excel files in folder, excluding temp files."""
    return [
        f for f in folder_path.iterdir()
        if f.is_file()
           and f.suffix.lower() == ".xlsx"
           and not f.name.startswith('~$')
    ]


def _select_worksheet(wb, sheet: Union[str, int], file_name: str):
    """
    Select worksheet from workbook by index or name.

    Returns:
        Worksheet or None if not found.
    """
    if isinstance(sheet, int):
        if sheet < 0 or sheet >= len(wb.sheetnames):
            logger.warning(f"Skipping {file_name} - sheet index {sheet} out of range.")
            return None
        return wb[wb.sheetnames[sheet]]
    else:
        if sheet not in wb.sheetnames:
            logger.warning(f"Skipping {file_name} - sheet '{sheet}' not found.")
            return None
        return wb[sheet]


def _extract_rows_from_worksheet(
    ws,
    current_filename: str,
    params: Union[ExtractionParams, DynamicExtractionParams],  # Updated
    left_idx: int,
    field_idx: int,
    value_idx: int
) -> FileExtractionResult:
    """Extract field mappings from a single worksheet."""
    file_data: MappingDict = defaultdict(list)
    file_filename_info: MappingDict = defaultdict(list)
    file_expression_filenames: MappingDict = defaultdict(list)

    max_col_idx = max(left_idx, field_idx, value_idx)

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= params.header_rows_to_skip:
            continue

        if max_col_idx - 1 >= len(row):
            continue

        left_val = _normalize_value(row[left_idx - 1], params.strip_key_parts)
        field_val = _normalize_value(row[field_idx - 1], params.strip_key_parts)
        value_val = _normalize_value(row[value_idx - 1], params.strip_key_parts)

        # Skip if key columns are empty
        if not left_val or not field_val:
            continue

        # Skip invalid values
        if _is_invalid_value(value_val):
            continue

        key = (left_val, field_val)

        # Add expression
        file_data[key].append(value_val)

        # Track unique filenames for this key
        if current_filename not in file_filename_info[key]:
            file_filename_info[key].append(current_filename)

        # Track filename for each expression (maintains order)
        file_expression_filenames[key].append(current_filename)

    return FileExtractionResult(
        mappings=dict(file_data),
        filename_info=dict(file_filename_info),
        expression_filenames=dict(file_expression_filenames)
    )


def _process_single_file(
        file_path: Path,
        params: ExtractionParams,
        left_idx: int,
        field_idx: int,
        value_idx: int
) -> Optional[FileExtractionResult]:
    """
    Process a single Excel file and extract field mappings.

    Returns:
        FileExtractionResult or None if file cannot be processed.
    """
    current_filename = file_path.stem

    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except (InvalidFileException, OSError) as e:
        logger.warning(f"Cannot open file '{file_path.name}': {e}")
        return None

    try:
        ws = _select_worksheet(wb, params.sheet, file_path.name)
        if ws is None:
            return None

        return _extract_rows_from_worksheet(
            ws, current_filename, params, left_idx, field_idx, value_idx
        )
    finally:
        wb.close()


def _merge_file_result(
        result: MappingDict,
        filename_info: MappingDict,
        expression_filenames: MappingDict,
        file_result: FileExtractionResult
) -> None:
    """Merge single file results into global results (modifies in-place)."""
    # Merge mappings
    for key, values in file_result.mappings.items():
        if key not in result:
            result[key] = []
        result[key].extend(values)

    # Merge filename info (unique filenames per key)
    for key, filenames in file_result.filename_info.items():
        if key not in filename_info:
            filename_info[key] = []
        for fname in filenames:
            if fname not in filename_info[key]:
                filename_info[key].append(fname)

    # Merge expression filenames (one filename per expression, in order)
    for key, filenames in file_result.expression_filenames.items():
        if key not in expression_filenames:
            expression_filenames[key] = []
        expression_filenames[key].extend(filenames)


def _validate_result_consistency(
        result: MappingDict,
        expression_filenames: MappingDict
) -> None:
    """Validate that expressions and filenames are properly aligned."""
    for key in result:
        expr_count = len(result[key])
        fname_count = len(expression_filenames.get(key, []))
        if expr_count != fname_count:
            logger.error(
                f"Data inconsistency for key {key}: "
                f"{expr_count} expressions vs {fname_count} filenames"
            )


def _try_get_from_cache(
        cache: Optional[ExcelCache],
        file_path: Path,
        extraction_params: dict,
        stats: CacheStats
) -> Optional[FileExtractionResult]:
    """Attempt to retrieve file result from cache."""
    if cache is None:
        return None

    cached_data = cache.get_cached_extraction_with_filenames(file_path, extraction_params)

    if cached_data and isinstance(cached_data, tuple) and len(cached_data) == 3:
        stats.record_hit()
        return FileExtractionResult.from_tuple(cached_data)

    stats.record_miss()
    return None


def _save_to_cache(
        cache: Optional[ExcelCache],
        file_path: Path,
        extraction_params: dict,
        file_result: FileExtractionResult
) -> None:
    """Save extraction result to cache."""
    if cache is not None:
        cache.save_extraction_result_with_filenames(
            file_path,
            extraction_params,
            file_result.to_tuple()
        )

def _normalize_header_text(text) -> str:
    """
    Normalize header text for matching.
    Strips whitespace, converts to lowercase, and collapses multiple spaces.
    """
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _find_dynamic_value_column(
    ws,
    possible_value_idxs: List[int],
    normalized_match: str,
    header_search_rows: int
) -> Optional[int]:
    """
    Find the value column by searching for matching header text.

    Scans the first `header_search_rows` rows looking for a cell
    in one of the `possible_value_idxs` that matches the normalized text.

    Returns:
        1-based column index or None if not found.
    """
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx > header_search_rows:
            break
        for c_idx in possible_value_idxs:
            cell_val = (
                _normalize_header_text(row[c_idx - 1])
                if c_idx - 1 < len(row)
                else ""
            )
            if cell_val == normalized_match:
                return c_idx
    return None


def _process_single_file_dynamic(
    file_path: Path,
    params: DynamicExtractionParams,
    left_idx: int,
    field_idx: int,
    possible_value_idxs: List[int],
    normalized_match: str,
    possible_value_cols: List[str]
) -> Optional[FileExtractionResult]:
    """
    Process a single Excel file with dynamic column detection.

    Returns:
        FileExtractionResult or None if file cannot be processed.
    """
    current_filename = file_path.stem

    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except (InvalidFileException, OSError) as e:
        logger.warning(f"Cannot open file '{file_path.name}': {e}")
        return None

    try:
        ws = _select_worksheet(wb, params.sheet, file_path.name)
        if ws is None:
            return None

        # Find the dynamic value column
        value_idx = _find_dynamic_value_column(
            ws,
            possible_value_idxs,
            normalized_match,
            params.header_search_rows
        )

        if value_idx is None:
            logger.warning(
                f"Skipping {file_path.name} - "
                f"no column found with header '{params.match_text}'."
            )
            return None

        # Log which column was detected
        col_letter = possible_value_cols[possible_value_idxs.index(value_idx)]
        logger.debug(f"File {file_path.name}: detected value column = {col_letter}")

        # Reuse the shared extraction function
        return _extract_rows_from_worksheet(
            ws, current_filename, params, left_idx, field_idx, value_idx
        )
    finally:
        wb.close()


def extract_combined_field_mappings_from_folder(
        folder_path: str,
        left_col: Union[str, int] = "B",
        field_col: Union[str, int] = "C",
        value_col: Union[str, int] = "D",
        sheet: Union[str, int] = 0,
        header_rows_to_skip: int = 2,
        strip_key_parts: bool = True,
        cache: Optional[ExcelCache] = None,
        use_cache: bool = True
) -> ExtractionResult:
    """
    Extracts mappings from all Excel files in a folder.

    For each file:
      Key = (column B, column C)
      Value = list of column D entries aggregated across all files.

    IMPORTANT: Each expression is tracked with its source filename so we know
    which file each expression came from.

    Args:
        folder_path: Path to folder containing Excel files
        left_col: Column for left part of key (default "B" for Dim./Meas.)
        field_col: Column for field name (default "C" for Name)
        value_col: Column for mapping value (default "D" for Mapping)
        sheet: Sheet index or name (default 0)
        header_rows_to_skip: Number of header rows to skip (default 2)
        strip_key_parts: Whether to strip whitespace from keys (default True)
        cache: Optional cache instance to use for caching
        use_cache: Whether to use caching (default True)

    Returns:
        Tuple of three dictionaries:
        - mappings: (dim, field) tuples to lists of mapping expressions
        - filename_info: (dim, field) tuples to lists of unique filenames
        - expression_filenames: (dim, field) tuples to lists of filenames
    """
    # Build extraction parameters
    params = ExtractionParams(
        left_col=left_col,
        field_col=field_col,
        value_col=value_col,
        sheet=sheet,
        header_rows_to_skip=header_rows_to_skip,
        strip_key_parts=strip_key_parts
    )

    # Pre-compute column indices
    left_idx = _col_to_index(left_col)
    field_idx = _col_to_index(field_col)
    value_idx = _col_to_index(value_col)

    # Validate folder
    folder_obj = Path(folder_path)
    if not folder_obj.exists() or not folder_obj.is_dir():
        raise NotADirectoryError(f"Invalid folder path: {folder_path}")

    # Initialize cache if needed
    if cache is None and use_cache:
        cache = get_cache()
    effective_cache = cache if use_cache else None

    # Result dictionaries
    result: MappingDict = {}
    filename_info: MappingDict = {}
    expression_filenames: MappingDict = {}

    # Get Excel files
    excel_files = _get_excel_files(folder_obj)

    if not excel_files:
        logger.info(f"No Excel files found in the folder: {folder_path}")
        return result, filename_info, expression_filenames

    logger.info(f"Processing {len(excel_files)} Excel file(s) from folder: {folder_path}")

    # Cache statistics
    stats = CacheStats()
    extraction_params_dict = params.to_dict()

    for file_path in excel_files:
        try:
            # Try cache first
            file_result = _try_get_from_cache(
                effective_cache, file_path, extraction_params_dict, stats
            )

            # Process file if not cached
            if file_result is None:
                stats.record_processed()
                file_result = _process_single_file(
                    file_path, params, left_idx, field_idx, value_idx
                )

                if file_result is None:
                    continue

                # Save to cache
                _save_to_cache(
                    effective_cache, file_path, extraction_params_dict, file_result
                )

                logger.debug(
                    f"Processed: {file_path.name} ({len(file_result.mappings)} mappings)"
                )

            # Merge into global results
            _merge_file_result(result, filename_info, expression_filenames, file_result)

        except (OSError, KeyError) as e:
            logger.warning(f"Error reading file '{file_path.name}': {e}")
            continue
        except Exception as e:
            logger.error(
                f"Unexpected error processing '{file_path.name}': {e}",
                exc_info=True
            )
            continue

    # Validate consistency
    _validate_result_consistency(result, expression_filenames)

    # Log cache stats
    if use_cache and cache:
        stats.log_stats()

    logger.info(
        f"Completed processing {len(excel_files)} file(s). "
        f"Found {len(result)} unique keys."
    )

    return result, filename_info, expression_filenames


def extract_dynamic_mapping_column_from_folder_for_pi(
    folder_path: str,
    left_col: Union[str, int] = "B",
    field_col: Union[str, int] = "C",
    possible_value_cols: Optional[List[str]] = None,
    sheet: Union[str, int] = 0,
    header_rows_to_skip: int = 2,
    strip_key_parts: bool = True,
    header_search_rows: int = 5,
    match_text: str = "Mapping (from layout)",
    cache: Optional[ExcelCache] = None,
    use_cache: bool = True
) -> ExtractionResult:
    """
    Extracts mappings from Excel files in a folder where the 'value' column
    must be dynamically determined by header content.

    The correct value column is identified by scanning the first
    `header_search_rows` rows for a cell matching `match_text`
    (case-insensitive, trimmed, normalized spaces) in one of the
    `possible_value_cols`.

    Args:
        folder_path: Path to folder containing Excel files.
        left_col: Column for left part of key (default "B")
        field_col: Column for field name (default "C")
        possible_value_cols: Columns to check for mapping header (default D–H)
        sheet: Sheet index or name (default 0)
        header_rows_to_skip: Number of header rows to skip for data (default 2)
        strip_key_parts: Whether to strip whitespace from keys (default True)
        header_search_rows: How many top rows to scan for header label (default 5)
        match_text: Text used to detect the value column
        cache: Optional cache instance to use for caching
        use_cache: Whether to use caching (default True)

    Returns:
        Tuple of three dictionaries:
        - mappings: (dim, field) tuples to lists of mapping expressions
        - filename_info: (dim, field) tuples to lists of unique filenames
        - expression_filenames: (dim, field) tuples to lists of filenames
    """
    if possible_value_cols is None:
        possible_value_cols = ["D", "E", "F", "G", "H"]

    params = DynamicExtractionParams(
        left_col=left_col,
        field_col=field_col,
        possible_value_cols=tuple(possible_value_cols),
        sheet=sheet,
        header_rows_to_skip=header_rows_to_skip,
        strip_key_parts=strip_key_parts,
        header_search_rows=header_search_rows,
        match_text=match_text
    )

    left_idx = _col_to_index(left_col)
    field_idx = _col_to_index(field_col)
    possible_value_idxs = [_col_to_index(c) for c in possible_value_cols]
    normalized_match = _normalize_header_text(match_text)

    folder_obj = Path(folder_path)
    if not folder_obj.exists() or not folder_obj.is_dir():
        raise NotADirectoryError(f"Invalid folder path: {folder_path}")

    if cache is None and use_cache:
        cache = get_cache()
    effective_cache = cache if use_cache else None

    result: MappingDict = {}
    filename_info: MappingDict = {}
    expression_filenames: MappingDict = {}

    excel_files = _get_excel_files(folder_obj)

    if not excel_files:
        logger.info(f"No Excel files found in folder: {folder_path}")
        return result, filename_info, expression_filenames

    logger.info(f"Processing {len(excel_files)} Excel file(s) from folder: {folder_path}")

    stats = CacheStats()
    extraction_params_dict = params.to_dict()

    for file_path in excel_files:
        try:
            file_result = _try_get_from_cache(
                effective_cache, file_path, extraction_params_dict, stats
            )

            if file_result is None:
                stats.record_processed()
                file_result = _process_single_file_dynamic(
                    file_path,
                    params,
                    left_idx,
                    field_idx,
                    possible_value_idxs,
                    normalized_match,
                    possible_value_cols
                )

                if file_result is None:
                    continue

                _save_to_cache(
                    effective_cache, file_path, extraction_params_dict, file_result
                )

                logger.debug(
                    f"Processed: {file_path.name} ({len(file_result.mappings)} mappings)"
                )

            _merge_file_result(result, filename_info, expression_filenames, file_result)

        except (OSError, KeyError) as e:
            logger.warning(f"Error reading file '{file_path.name}': {e}")
            continue
        except Exception as e:
            logger.error(
                f"Unexpected error processing '{file_path.name}': {e}",
                exc_info=True
            )
            continue

    _validate_result_consistency(result, expression_filenames)

    if use_cache and effective_cache:
        stats.log_stats()

    logger.info(
        f"Completed processing {len(excel_files)} file(s). "
        f"Found {len(result)} unique keys."
    )

    return result, filename_info, expression_filenames


# # folder_path = r'C:\Users\aditya.prasad\OneDrive - Mobileum\Documents\OneDrive - Mobileum\Templates Test - Templates\RA\UC\MSC\Ericsson\LdRules'
# folder_path = r'C:\Users\aditya.prasad\OneDrive - Mobileum\Documents\OneDrive - Mobileum\Template Hierarchy Structure - Templates\RA\PI\HLR\Nokia\LdRules'
# result, filename_info, expression_filenames = extract_dynamic_mapping_column_from_folder_for_pi(folder_path)
# print(result)
# print('----------------\n\n')
# print(filename_info)
# print('----------------\n\n')
# print(expression_filenames)