"""
Layout file extractor for reading and formatting Layout Excel files.
Supports operator extraction from filenames and Excel content reading.
"""
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
from openpyxl import load_workbook


def extract_operator_from_layout_filename(
    filename: str,
    source: str,
    vendor: str
) -> str:
    """
    Extract operator name from Layout filename by removing known parts.

    Logic:
    1. Remove 'Layout' prefix (case-insensitive)
    2. Remove source (case-insensitive)
    3. Remove vendor (case-insensitive)
    4. Remove numeric values
    5. Remove common separators and clean up
    6. What remains is the operator name

    Args:
        filename: The Excel filename (without extension)
        source: The source name from folder structure
        vendor: The vendor name from folder structure

    Returns:
        Extracted operator name or "Unknown" if not found
    """
    if not filename:
        return "Unknown"

    # Start with the filename
    remaining = filename

    # Create patterns to remove (case-insensitive)
    patterns_to_remove = [
        r'(?i)^Layout[_\-\s]*',  # Layout at start
        r'(?i)[_\-\s]*Layout$',  # Layout at end
        r'(?i)[_\-\s]*Layout[_\-\s]*',  # Layout anywhere
    ]

    # Remove Layout
    for pattern in patterns_to_remove:
        remaining = re.sub(pattern, '_', remaining)

    # Split by common separators
    parts = re.split(r'[_\-\s]+', remaining)

    # Filter out parts that match source, vendor, or are numeric
    filtered_parts = []
    source_lower = source.lower()
    vendor_lower = vendor.lower()

    for part in parts:
        part_lower = part.lower().strip()

        # Skip empty parts
        if not part_lower:
            continue

        # Skip if it matches source (exact or partial)
        if part_lower == source_lower or source_lower in part_lower or part_lower in source_lower:
            continue

        # Skip if it matches vendor (exact or partial)
        if part_lower == vendor_lower or vendor_lower in part_lower or part_lower in vendor_lower:
            continue

        # Skip if it's purely numeric
        if part_lower.isdigit():
            continue

        # Skip common non-operator terms
        skip_terms = {'layout', 'template', 'v1', 'v2', 'v3', 'final', 'new', 'old', 'copy', 'format', 'decoder'}
        if part_lower in skip_terms:
            continue

        filtered_parts.append(part)

    # Join remaining parts
    if filtered_parts:
        operator = '_'.join(filtered_parts)
        return operator if operator else "Unknown"

    return "Unknown"


def get_layout_folder_path(
    root_folder: str,
    domain: str,
    module: str,
    source: str,
    vendor: str
) -> Optional[Path]:
    """
    Get the Layout folder path for given hierarchy.

    Args:
        root_folder: Root folder path
        domain: Domain name
        module: Module name
        source: Source name
        vendor: Vendor name

    Returns:
        Path to Layout folder or None if not found
    """
    layout_path = Path(root_folder) / domain / module / source / vendor / "Layout"
    if layout_path.exists() and layout_path.is_dir():
        return layout_path
    return None


def scan_layout_files(layout_folder: Path) -> List[Path]:
    """
    Scan Layout folder for Excel files.

    Args:
        layout_folder: Path to Layout folder

    Returns:
        List of Excel file paths
    """
    if not layout_folder or not layout_folder.exists():
        return []

    excel_files = [
        f for f in layout_folder.iterdir()
        if f.is_file() and f.suffix.lower() == ".xlsx" and not f.name.startswith('~$')
    ]
    return sorted(excel_files, key=lambda x: x.name)


def get_available_operators(
    layout_folder: Path,
    source: str,
    vendor: str
) -> Dict[str, Path]:
    """
    Get available operators from Layout folder.

    Args:
        layout_folder: Path to Layout folder
        source: Source name for operator extraction
        vendor: Vendor name for operator extraction

    Returns:
        Dictionary mapping operator names to file paths
    """
    operators = {}
    excel_files = scan_layout_files(layout_folder)

    for file_path in excel_files:
        filename = file_path.stem  # filename without extension
        operator = extract_operator_from_layout_filename(filename, source, vendor)
        operators[operator] = file_path

    return operators


def read_excel_content(
    file_path: Path,
    max_rows: Optional[int] = None
) -> Dict[str, Any]:
    """
    Read Excel file content including all sheets.

    Args:
        file_path: Path to Excel file
        max_rows: Maximum rows to read (None for all)

    Returns:
        Dictionary with file info and sheet data
    """
    result = {
        'filename': file_path.name,
        'filepath': str(file_path),
        'sheets': [],
        'total_rows': 0,
        'error': None
    }

    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_data = {
                'name': sheet_name,
                'headers': [],
                'rows': [],
                'row_count': 0,
                'col_count': 0
            }

            rows_read = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                # Skip completely empty rows
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue

                # First non-empty row is headers
                if i == 0 or (not sheet_data['headers'] and rows_read == 0):
                    sheet_data['headers'] = [str(cell) if cell is not None else '' for cell in row]
                    sheet_data['col_count'] = len(sheet_data['headers'])
                else:
                    # Data rows
                    row_data = [str(cell) if cell is not None else '' for cell in row]
                    sheet_data['rows'].append(row_data)
                    rows_read += 1

                    if max_rows and rows_read >= max_rows:
                        break

            sheet_data['row_count'] = len(sheet_data['rows'])
            result['sheets'].append(sheet_data)
            result['total_rows'] += sheet_data['row_count']

        wb.close()

    except Exception as e:
        result['error'] = str(e)

    return result


def format_excel_as_markdown_table(
    sheet_data: Dict[str, Any],
    max_col_width: int = 30
) -> str:
    """
    Format sheet data as markdown table.

    Args:
        sheet_data: Sheet data dictionary
        max_col_width: Maximum column width for truncation

    Returns:
        Markdown formatted table string
    """
    if not sheet_data['headers']:
        return "*Empty sheet*"

    def truncate(text: str, max_len: int) -> str:
        """Truncate text with ellipsis if too long."""
        text = str(text).replace('\n', ' ').replace('\r', '')
        if len(text) > max_len:
            return text[:max_len-3] + '...'
        return text

    # Build header row
    headers = [truncate(h, max_col_width) for h in sheet_data['headers']]
    header_row = '| ' + ' | '.join(headers) + ' |'

    # Build separator row
    separator = '| ' + ' | '.join(['---'] * len(headers)) + ' |'

    # Build data rows
    data_rows = []
    for row in sheet_data['rows']:
        # Pad row if needed
        padded_row = row + [''] * (len(headers) - len(row))
        formatted_row = [truncate(cell, max_col_width) for cell in padded_row[:len(headers)]]
        data_rows.append('| ' + ' | '.join(formatted_row) + ' |')

    # Combine
    table = '\n'.join([header_row, separator] + data_rows)
    return table


def format_layout_content_html(
    content: Dict[str, Any],
    max_col_width: int = 50
) -> str:
    """
    Format layout content as HTML for scrollable display.

    Args:
        content: Excel content dictionary
        max_col_width: Maximum column width

    Returns:
        HTML formatted string
    """
    if content.get('error'):
        return f"<p style='color: red;'>Error reading file: {content['error']}</p>"

    html_parts = []

    for sheet in content['sheets']:
        # Sheet header
        html_parts.append(f"<h4>Sheet: {sheet['name']} ({sheet['row_count']} rows × {sheet['col_count']} columns)</h4>")

        if not sheet['headers']:
            html_parts.append("<p><em>Empty sheet</em></p>")
            continue

        # Build table
        html_parts.append("<div style='overflow-x: auto;'>")
        html_parts.append("<table style='border-collapse: collapse; width: 100%; font-size: 12px;'>")

        # Header row
        html_parts.append("<thead><tr style='background-color: #f0f0f0;'>")
        for header in sheet['headers']:
            html_parts.append(f"<th style='border: 1px solid #ddd; padding: 8px; text-align: left;'>{header}</th>")
        html_parts.append("</tr></thead>")

        # Data rows
        html_parts.append("<tbody>")
        for i, row in enumerate(sheet['rows']):
            bg_color = '#ffffff' if i % 2 == 0 else '#f9f9f9'
            html_parts.append(f"<tr style='background-color: {bg_color};'>")
            for j, cell in enumerate(row):
                if j < len(sheet['headers']):
                    html_parts.append(f"<td style='border: 1px solid #ddd; padding: 8px;'>{cell}</td>")
            # Pad if row has fewer columns
            for _ in range(len(sheet['headers']) - len(row)):
                html_parts.append("<td style='border: 1px solid #ddd; padding: 8px;'></td>")
            html_parts.append("</tr>")
        html_parts.append("</tbody>")

        html_parts.append("</table>")
        html_parts.append("</div>")
        html_parts.append("<br/>")

    return '\n'.join(html_parts)


def format_layout_content_text(content: Dict[str, Any]) -> str:
    """
    Format layout content as plain text for copying.

    Args:
        content: Excel content dictionary

    Returns:
        Tab-separated text string
    """
    if content.get('error'):
        return f"Error reading file: {content['error']}"

    text_parts = []

    for sheet in content['sheets']:
        text_parts.append(f"=== Sheet: {sheet['name']} ===")
        text_parts.append("")

        if not sheet['headers']:
            text_parts.append("(Empty sheet)")
            text_parts.append("")
            continue

        # Header row
        text_parts.append('\t'.join(sheet['headers']))

        # Data rows
        for row in sheet['rows']:
            text_parts.append('\t'.join(row))

        text_parts.append("")

    return '\n'.join(text_parts)
